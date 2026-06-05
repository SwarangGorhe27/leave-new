"""
Swipe Log Export Service - CSV/XLSX/PDF export functionality.

Handles data retrieval, formatting, and file generation for swipe log exports.
"""

import csv
import io
import logging
from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone

from apps.attendance.models import SwipeLogExportJob, PunchLog
from apps.attendance.selectors.swipe_logs.export_import_selectors import get_filtered_swipe_logs

logger = logging.getLogger(__name__)


class SwipeLogExportService:
    """Service for exporting swipe logs to various formats."""

    def __init__(self):
        """Initialize export service."""
        pass

    def _build_filters(self, export_job: SwipeLogExportJob) -> dict:
        return {
            "from_date": export_job.from_date,
            "to_date": export_job.to_date,
            "employee_ids": export_job.employee_ids,
            "department_ids": export_job.department_ids,
            "punch_type": export_job.punch_types[0] if export_job.punch_types else None,
        }

    def export_to_csv(self, export_job_id: str) -> str:
        """
        Export swipe logs to CSV format.

        Args:
            export_job_id: Export job UUID

        Returns:
            CSV content as string
        """
        try:
            export_job = SwipeLogExportJob.objects.get(id=export_job_id)
            logger.info(f"Exporting to CSV: {export_job_id}")

            # Build filters
            filters = self._build_filters(export_job)

            # Get filtered data
            queryset = get_filtered_swipe_logs(export_job.company_id, filters)

            # Fetch data with related objects
            from apps.attendance.utils.employee_relations import defer_employment_team

            punch_logs = defer_employment_team(
                queryset.select_related(
                    "employee",
                    "employee__employment_details",
                    "employee__employment_details__department",
                    "device",
                )
            ).values_list(
                "id",
                "employee__employee_code",
                "employee__first_name",
                "employee__last_name",
                "employee__employment_details__department__name",
                "punch_time",
                "punch_type",
                "punch_source",
                "device__device_name",
                "face_verified",
                "is_trusted",
                "latitude",
                "longitude",
                "is_within_geofence"
            )

            # Build CSV content
            output = io.StringIO()
            writer = csv.writer(output)

            # Write header
            headers = [
                "Punch ID",
                "Employee Code",
                "First Name",
                "Last Name",
                "Department",
                "Punch Time",
                "Punch Type",
                "Punch Source",
                "Device Name",
            ]

            if export_job.include_verification_details:
                headers.extend(["Face Verified", "Is Trusted"])

            if export_job.include_geofence_details:
                headers.extend(["Latitude", "Longitude", "Within Geofence"])

            writer.writerow(headers)

            # Write data rows
            processed = 0
            for row in punch_logs:
                csv_row = list(row[:9])  # Base columns
                
                if export_job.include_verification_details:
                    csv_row.extend([row[9], row[10]])  # face_verified, is_trusted
                
                if export_job.include_geofence_details:
                    csv_row.extend([row[11], row[12], row[13]])  # lat, lon, geofence

                writer.writerow(csv_row)
                processed += 1

            # Update processed count
            export_job.processed_records = processed
            export_job.save(update_fields=["processed_records"])

            content = output.getvalue()
            logger.info(f"CSV export completed: {export_job_id} ({processed} records)")
            return content

        except Exception as e:
            logger.error(f"CSV export failed: {export_job_id} - {str(e)}", exc_info=True)
            raise

    def export_to_xlsx(self, export_job_id: str) -> bytes:
        """
        Export swipe logs to XLSX format.

        Args:
            export_job_id: Export job UUID

        Returns:
            XLSX content as bytes
        """
        try:
            export_job = SwipeLogExportJob.objects.get(id=export_job_id)
            logger.info(f"Exporting to XLSX: {export_job_id}")

            # Build filters
            filters = self._build_filters(export_job)

            # Get filtered data
            queryset = get_filtered_swipe_logs(export_job.company_id, filters)

            # Fetch data
            from apps.attendance.utils.employee_relations import defer_employment_team

            punch_logs = defer_employment_team(
                queryset.select_related(
                    "employee",
                    "employee__employment_details",
                    "employee__employment_details__department",
                    "device",
                )
            ).values_list(
                "id",
                "employee__employee_code",
                "employee__first_name",
                "employee__last_name",
                "employee__employment_details__department__name",
                "punch_time",
                "punch_type",
                "punch_source",
                "device__device_name",
                "face_verified",
                "is_trusted",
                "latitude",
                "longitude",
                "is_within_geofence"
            )

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Swipe Logs"

            # Define styles
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            # Build headers
            headers = [
                "Punch ID",
                "Employee Code",
                "First Name",
                "Last Name",
                "Department",
                "Punch Time",
                "Punch Type",
                "Punch Source",
                "Device Name",
            ]

            if export_job.include_verification_details:
                headers.extend(["Face Verified", "Is Trusted"])

            if export_job.include_geofence_details:
                headers.extend(["Latitude", "Longitude", "Within Geofence"])

            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border

            # Write data
            processed = 0
            for row_idx, row in enumerate(punch_logs, 2):
                xlsx_row = list(row[:9])  # Base columns
                
                if export_job.include_verification_details:
                    xlsx_row.extend([row[9], row[10]])
                
                if export_job.include_geofence_details:
                    xlsx_row.extend([row[11], row[12], row[13]])

                for col_idx, value in enumerate(xlsx_row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.alignment = center_alignment
                    cell.border = border

                processed += 1

            # Auto-adjust column widths
            for col_idx, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_idx)
                ws.column_dimensions[column_letter].width = max(12, len(str(header)) + 2)

            # Update processed count
            export_job.processed_records = processed
            export_job.save(update_fields=["processed_records"])

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            content = output.getvalue()

            logger.info(f"XLSX export completed: {export_job_id} ({processed} records)")
            return content

        except Exception as e:
            logger.error(f"XLSX export failed: {export_job_id} - {str(e)}", exc_info=True)
            raise

    def export_to_pdf(self, export_job_id: str) -> bytes:
        """
        Export swipe logs to PDF format using ReportLab.

        Args:
            export_job_id: Export job UUID

        Returns:
            PDF content as bytes
        """
        try:
            export_job = SwipeLogExportJob.objects.get(id=export_job_id)
            logger.info(f"Exporting to PDF: {export_job_id}")

            # Build filters
            filters = self._build_filters(export_job)

            # Get filtered data
            queryset = get_filtered_swipe_logs(export_job.company_id, filters)

            # Fetch data
            from apps.attendance.utils.employee_relations import defer_employment_team

            punch_logs = defer_employment_team(
                queryset.select_related(
                    "employee",
                    "employee__employment_details",
                    "employee__employment_details__department",
                    "device",
                )
            ).values_list(
                "id",
                "employee__employee_code",
                "employee__first_name",
                "employee__last_name",
                "employee__employment_details__department__name",
                "punch_time",
                "punch_type",
                "punch_source",
                "device__device_name",
                "face_verified",
                "is_trusted",
                "latitude",
                "longitude",
                "is_within_geofence"
            )

            # Convert to list for processing
            data_list = list(punch_logs)
            processed = len(data_list)

            # Generate PDF using xhtml2pdf and validate the output with PyMuPDF
            pdf_bytes = self._generate_pdf_xhtml2pdf(export_job, data_list)
            pdf_bytes = self._validate_pdf_bytes(pdf_bytes)

            # Update processed count
            export_job.processed_records = processed
            export_job.save(update_fields=["processed_records"])

            logger.info(f"PDF export completed: {export_job_id} ({processed} records)")
            return pdf_bytes

        except Exception as e:
            logger.error(f"PDF export failed: {export_job_id} - {str(e)}", exc_info=True)
            raise

    def _generate_pdf_xhtml2pdf(self, export_job: SwipeLogExportJob, data_list: List) -> bytes:
        """
        Generate PDF using xhtml2pdf (HTML to PDF) and PyMuPDF for post-processing if needed.

        Args:
            export_job: Export job instance
            data_list: List of punch log data tuples

        Returns:
            PDF content as bytes
        """
        try:
            from xhtml2pdf import pisa
        except ModuleNotFoundError as exc:
            raise ImportError(
                "PDF export requires the optional dependency 'xhtml2pdf==0.2.17'. "
                "Install it in the active virtual environment to enable PDF export."
            ) from exc

        from io import BytesIO
        from django.utils.html import escape

        # Build table headers
        headers = [
            "Punch ID",
            "Employee Code",
            "First Name",
            "Last Name",
            "Department",
            "Punch Time",
            "Punch Type",
            "Punch Source",
            "Device Name",
        ]

        if export_job.include_verification_details:
            headers.extend(["Face Verified", "Is Trusted"])

        if export_job.include_geofence_details:
            headers.extend(["Latitude", "Longitude", "Within Geofence"])

        # Build HTML content
        html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; font-size: 10pt; }}
                h1 {{ color: #333; text-align: center; }}
                .info {{ color: #666; font-size: 9pt; margin-bottom: 10px; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
                th, td {{ border: 1px solid #ccc; padding: 4px 6px; text-align: center; }}
                th {{ background: #4472C4; color: #fff; }}
                tr:nth-child(even) {{ background: #f9f9f9; }}
            </style>
        </head>
        <body>
            <h1>Swipe Log Export Report</h1>
            <div class="info">
                <b>Generated:</b> {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>
                <b>From Date:</b> {escape(str(export_job.from_date) if export_job.from_date else 'N/A')}<br/>
                <b>To Date:</b> {escape(str(export_job.to_date) if export_job.to_date else 'N/A')}<br/>
                <b>Total Records:</b> {escape(str(export_job.total_records))}
            </div>
            <table>
                <tr>
                    {''.join(f'<th>{escape(h)}</th>' for h in headers)}
                </tr>
        """
        for row in data_list:
            html += "<tr>"
            row_data = [str(cell) if cell is not None else "" for cell in row[:9]]
            if export_job.include_verification_details:
                row_data.extend([str(row[9]) if row[9] is not None else "",
                                str(row[10]) if row[10] is not None else ""])
            if export_job.include_geofence_details:
                row_data.extend([str(row[11]) if row[11] is not None else "",
                                str(row[12]) if row[12] is not None else "",
                                str(row[13]) if row[13] is not None else ""])
            for cell in row_data:
                html += f"<td>{escape(cell)}</td>"
            html += "</tr>"
        html += """
            </table>
        </body>
        </html>
        """

        # Generate PDF from HTML
        pdf_buffer = BytesIO()
        pisa_status = pisa.CreatePDF(html, dest=pdf_buffer)
        if pisa_status.err:
            raise Exception("Error generating PDF with xhtml2pdf")
        pdf_buffer.seek(0)
        return pdf_buffer.getvalue()

    def _validate_pdf_bytes(self, pdf_bytes: bytes) -> bytes:
        """
        Validate generated PDF bytes with PyMuPDF and normalize the stream.
        """
        try:
            import fitz
        except ModuleNotFoundError as exc:
            raise ImportError(
                "PDF validation requires the optional dependency 'pymupdf==1.25.5'. "
                "Install it in the active virtual environment to enable PDF export."
            ) from exc

        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        try:
            if doc.page_count == 0:
                raise Exception("Generated PDF contains no pages")
            normalized = doc.tobytes(deflate=True, garbage=4)
        finally:
            doc.close()

        check = fitz.open(stream=normalized, filetype="pdf")
        try:
            if check.page_count == 0:
                raise Exception("Validated PDF contains no pages")
        finally:
            check.close()

        return normalized
