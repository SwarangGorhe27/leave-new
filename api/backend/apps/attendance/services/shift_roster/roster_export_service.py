"""Service for Shift Roster Export (CSV/XLSX)."""

import logging
import io
from typing import Dict, Any, Optional, List
from datetime import date
from calendar import monthrange

try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

import csv

from django.http import HttpResponse

from apps.attendance.models import EmployeeShiftRoster
from apps.employees.models import Company

logger = logging.getLogger(__name__)


class RosterExportService:
    """
    Service for exporting shift rosters in various formats.
    
    Handles:
    - CSV export
    - XLSX export
    - Data formatting
    - File generation
    """

    @staticmethod
    def get_roster_export_data(
        company: Company,
        month: int,
        year: int,
        department_id: Optional[str] = None,
        designation_id: Optional[str] = None,
    ) -> List[Dict[str, Any]]:
        """
        Prepare roster data for export.
        
        Args:
            company: Company instance
            month: Month number
            year: Year
            department_id: Filter by department (optional)
            designation_id: Filter by designation (optional)
            
        Returns:
            List of roster records for export
        """
        try:
            logger.info(
                f"Preparing roster export for {month}/{year}",
                extra={
                    "company_id": str(company.id),
                    "month": month,
                    "year": year,
                },
            )

            first_day = date(year, month, 1)
            last_day = date(year, month, monthrange(year, month)[1])

            queryset = EmployeeShiftRoster.objects.filter(
                company=company,
                roster_date__gte=first_day,
                roster_date__lte=last_day,
                deleted_at__isnull=True,
            ).select_related(
                "employee",
                "shift",
                "employee__department",
                "employee__designation",
            ).order_by("roster_date", "employee__employee_code")

            if department_id:
                queryset = queryset.filter(employee__department_id=department_id)

            if designation_id:
                queryset = queryset.filter(employee__designation_id=designation_id)

            # Prepare export data
            export_data = []
            for roster in queryset:
                export_data.append({
                    "employee_code": roster.employee.employee_code,
                    "employee_name": roster.employee.get_full_name(),
                    "department": roster.employee.department.name if roster.employee.department else "N/A",
                    "designation": roster.employee.designation.name if hasattr(roster.employee, "designation") and roster.employee.designation else "N/A",
                    "roster_date": roster.roster_date.strftime("%Y-%m-%d"),
                    "shift_code": roster.shift.code,
                    "shift_name": roster.shift.name,
                    "start_time": roster.shift.start_time.strftime("%H:%M"),
                    "end_time": roster.shift.end_time.strftime("%H:%M"),
                    "is_week_off": "Yes" if roster.is_week_off else "No",
                    "override_reason": roster.override_reason or "",
                })

            logger.debug(
                f"Roster export data prepared: {len(export_data)} records",
                extra={
                    "company_id": str(company.id),
                    "record_count": len(export_data),
                },
            )

            return export_data

        except Exception as e:
            logger.error(
                f"Error preparing roster export: {str(e)}",
                extra={"company_id": str(company.id)},
                exc_info=True,
            )
            raise

    @staticmethod
    def export_to_csv(
        company: Company,
        month: int,
        year: int,
        department_id: Optional[str] = None,
        designation_id: Optional[str] = None,
    ) -> HttpResponse:
        """
        Export roster data to CSV format.
        
        Args:
            company: Company instance
            month: Month number
            year: Year
            department_id: Filter by department (optional)
            designation_id: Filter by designation (optional)
            
        Returns:
            HttpResponse with CSV file
        """
        try:
            logger.info(
                f"Exporting roster to CSV for {month}/{year}",
                extra={
                    "company_id": str(company.id),
                    "month": month,
                    "year": year,
                },
            )

            # Get export data
            export_data = RosterExportService.get_roster_export_data(
                company=company,
                month=month,
                year=year,
                department_id=department_id,
                designation_id=designation_id,
            )

            # Create CSV response
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = (
                f'attachment; filename="roster_{year}_{month:02d}.csv"'
            )

            # Write CSV
            writer = csv.DictWriter(
                response,
                fieldnames=[
                    "employee_code",
                    "employee_name",
                    "department",
                    "designation",
                    "roster_date",
                    "shift_code",
                    "shift_name",
                    "start_time",
                    "end_time",
                    "is_week_off",
                    "override_reason",
                ],
            )
            writer.writeheader()
            writer.writerows(export_data)

            logger.info(
                f"CSV export completed for {month}/{year}",
                extra={
                    "company_id": str(company.id),
                    "record_count": len(export_data),
                },
            )

            return response

        except Exception as e:
            logger.error(
                f"Error exporting roster to CSV: {str(e)}",
                extra={"company_id": str(company.id)},
                exc_info=True,
            )
            raise

    @staticmethod
    def export_to_xlsx(
        company: Company,
        month: int,
        year: int,
        department_id: Optional[str] = None,
        designation_id: Optional[str] = None,
    ) -> HttpResponse:
        """
        Export roster data to XLSX format.
        
        Args:
            company: Company instance
            month: Month number
            year: Year
            department_id: Filter by department (optional)
            designation_id: Filter by designation (optional)
            
        Returns:
            HttpResponse with XLSX file
        """
        if not XLSX_AVAILABLE:
            raise ImportError(
                "openpyxl is not installed. Please install it to use XLSX export."
            )

        try:
            logger.info(
                f"Exporting roster to XLSX for {month}/{year}",
                extra={
                    "company_id": str(company.id),
                    "month": month,
                    "year": year,
                },
            )

            # Get export data
            export_data = RosterExportService.get_roster_export_data(
                company=company,
                month=month,
                year=year,
                department_id=department_id,
                designation_id=designation_id,
            )

            # Create workbook
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Roster"

            # Header styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            # Write headers
            headers = [
                "Employee Code",
                "Employee Name",
                "Department",
                "Designation",
                "Roster Date",
                "Shift Code",
                "Shift Name",
                "Start Time",
                "End Time",
                "Week Off",
                "Override Reason",
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Write data
            for row_idx, data in enumerate(export_data, 2):
                ws.cell(row=row_idx, column=1).value = data["employee_code"]
                ws.cell(row=row_idx, column=2).value = data["employee_name"]
                ws.cell(row=row_idx, column=3).value = data["department"]
                ws.cell(row=row_idx, column=4).value = data["designation"]
                ws.cell(row=row_idx, column=5).value = data["roster_date"]
                ws.cell(row=row_idx, column=6).value = data["shift_code"]
                ws.cell(row=row_idx, column=7).value = data["shift_name"]
                ws.cell(row=row_idx, column=8).value = data["start_time"]
                ws.cell(row=row_idx, column=9).value = data["end_time"]
                ws.cell(row=row_idx, column=10).value = data["is_week_off"]
                ws.cell(row=row_idx, column=11).value = data["override_reason"]

            # Adjust column widths
            ws.column_dimensions["A"].width = 15
            ws.column_dimensions["B"].width = 25
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 20
            ws.column_dimensions["E"].width = 15
            ws.column_dimensions["F"].width = 15
            ws.column_dimensions["G"].width = 20
            ws.column_dimensions["H"].width = 12
            ws.column_dimensions["I"].width = 12
            ws.column_dimensions["J"].width = 12
            ws.column_dimensions["K"].width = 30

            # Create response
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = (
                f'attachment; filename="roster_{year}_{month:02d}.xlsx"'
            )

            # Save workbook
            wb.save(response)

            logger.info(
                f"XLSX export completed for {month}/{year}",
                extra={
                    "company_id": str(company.id),
                    "record_count": len(export_data),
                },
            )

            return response

        except Exception as e:
            logger.error(
                f"Error exporting roster to XLSX: {str(e)}",
                extra={"company_id": str(company.id)},
                exc_info=True,
            )
            raise
