from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django_tenants.utils import get_tenant_model, tenant_context
from openpyxl import load_workbook

from apps.employees.models.masters.education import Institution, University


DEFAULT_DOCUMENT_DIR = Path(__file__).resolve().parents[2] / "education_document"

UNIVERSITY_FILE = "University-ALL UNIVERSITIES.xlsx"
COLLEGE_FILE = "College-ALL COLLEGE.xlsx"
STANDALONE_FILE = "Standalone-Technical_Polytechnic.xlsx"


def clean_cell(value):
    if value is None:
        return None
    value = str(value).strip()
    return value or None


class Command(BaseCommand):
    help = (
        "Import education master lookup data from AISHE Excel files. "
        "Skips only Website and Year Of Establishment from the AISHE files."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--schema",
            type=str,
            required=True,
            help="Tenant schema name to import education data into.",
        )
        parser.add_argument(
            "--document-dir",
            type=str,
            default=str(DEFAULT_DOCUMENT_DIR),
            help="Directory containing the AISHE XLSX files.",
        )
        parser.add_argument(
            "--batch-size",
            type=int,
            default=2000,
            help="Rows to insert/update per batch.",
        )

    def handle(self, *args, **options):
        document_dir = Path(options["document_dir"])
        if not document_dir.exists():
            raise CommandError(f"Document directory not found: {document_dir}")

        batch_size = options["batch_size"]
        if batch_size <= 0:
            raise CommandError("--batch-size must be greater than 0")

        TenantModel = get_tenant_model()
        try:
            tenant = TenantModel.objects.get(schema_name=options["schema"])
        except TenantModel.DoesNotExist as exc:
            raise CommandError(
                f"Tenant with schema '{options['schema']}' not found"
            ) from exc

        with tenant_context(tenant):
            totals = self.import_files(document_dir, batch_size)

        self.stdout.write(
            self.style.SUCCESS(
                "Imported education masters: "
                f"{totals['universities']} universities, "
                f"{totals['institutions']} institutions "
                f"({totals['created']} created, {totals['updated']} updated)."
            )
        )

    def import_files(self, document_dir, batch_size):
        totals = {"universities": 0, "institutions": 0, "created": 0, "updated": 0}

        imported, created, updated = self.import_universities(
            document_dir / UNIVERSITY_FILE,
            batch_size,
        )
        totals["universities"] += imported
        totals["created"] += created
        totals["updated"] += updated

        university_by_code = {
            item.code: item for item in University.objects.only("id", "code")
        }
        university_type_by_code = {}

        for filename, institution_type in (
            (COLLEGE_FILE, Institution.InstitutionType.COLLEGE),
            (STANDALONE_FILE, Institution.InstitutionType.STANDALONE),
        ):
            imported, created, updated = self.import_institutions(
                document_dir / filename,
                institution_type,
                university_by_code,
                university_type_by_code,
                batch_size,
            )
            totals["institutions"] += imported
            totals["created"] += created
            totals["updated"] += updated

        self.update_university_types(university_type_by_code)

        return totals

    def get_rows(self, path, required_headers):
        if not path.exists():
            raise CommandError(f"Workbook not found: {path}")

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
        headers = [clean_cell(value) for value in header_row]
        header_indexes = {header: index for index, header in enumerate(headers) if header}
        missing = set(required_headers) - set(header_indexes)
        if missing:
            raise CommandError(
                f"{path.name} is missing required column(s): "
                + ", ".join(sorted(missing))
            )

        for row in ws.iter_rows(min_row=4, values_only=True):
            yield row, header_indexes

    def import_universities(self, path, batch_size):
        rows = []
        imported = created = updated = 0
        required_headers = ["Aishe Code", "Name", "State", "District", "Location"]
        for row, indexes in self.get_rows(path, required_headers):
            code = clean_cell(row[indexes["Aishe Code"]])
            label = clean_cell(row[indexes["Name"]])
            if not code or not label:
                continue
            rows.append(
                {
                    "code": code,
                    "label": label,
                    "state": clean_cell(row[indexes["State"]]),
                    "district": clean_cell(row[indexes["District"]]),
                    "location": clean_cell(row[indexes["Location"]]),
                    "is_active": True,
                }
            )
            if len(rows) >= batch_size:
                batch_created, batch_updated = self.upsert_universities(rows)
                imported += len(rows)
                created += batch_created
                updated += batch_updated
                rows = []

        if rows:
            batch_created, batch_updated = self.upsert_universities(rows)
            imported += len(rows)
            created += batch_created
            updated += batch_updated

        return imported, created, updated

    def import_institutions(
        self,
        path,
        institution_type,
        university_by_code,
        university_type_by_code,
        batch_size,
    ):
        rows = []
        imported = created = updated = 0
        required_headers = ["Aishe Code", "Name", "State", "District", "Location", "Manegement"]
        if institution_type == Institution.InstitutionType.COLLEGE:
            required_headers.extend(
                ["College Type", "University Aishe Code", "University Name", "University Type"]
            )
        else:
            required_headers.append("Standalone Type")

        for row, indexes in self.get_rows(path, required_headers):
            code = clean_cell(row[indexes["Aishe Code"]])
            label = clean_cell(row[indexes["Name"]])
            if not code or not label:
                continue
            university = None
            university_name = None
            university_type = None
            college_type = None
            standalone_type = None
            if institution_type == Institution.InstitutionType.COLLEGE:
                university_code = clean_cell(row[indexes["University Aishe Code"]])
                university = university_by_code.get(university_code)
                university_name = clean_cell(row[indexes["University Name"]])
                university_type = clean_cell(row[indexes["University Type"]])
                college_type = clean_cell(row[indexes["College Type"]])
                if university_code and university_type:
                    university_type_by_code[university_code] = university_type
            else:
                standalone_type = clean_cell(row[indexes["Standalone Type"]])
            rows.append(
                {
                    "code": code,
                    "label": label,
                    "institution_type": institution_type,
                    "university": university,
                    "state": clean_cell(row[indexes["State"]]),
                    "district": clean_cell(row[indexes["District"]]),
                    "location": clean_cell(row[indexes["Location"]]),
                    "college_type": college_type,
                    "standalone_type": standalone_type,
                    "management": clean_cell(row[indexes["Manegement"]]),
                    "university_name": university_name,
                    "university_type": university_type,
                    "is_active": True,
                }
            )
            if len(rows) >= batch_size:
                batch_created, batch_updated = self.upsert_institutions(rows)
                imported += len(rows)
                created += batch_created
                updated += batch_updated
                rows = []

        if rows:
            batch_created, batch_updated = self.upsert_institutions(rows)
            imported += len(rows)
            created += batch_created
            updated += batch_updated

        return imported, created, updated

    @transaction.atomic
    def upsert_universities(self, rows):
        existing = {
            item.code: item
            for item in University.objects.filter(code__in=[row["code"] for row in rows])
        }
        to_create = []
        to_update = []

        for row in rows:
            university = existing.get(row["code"])
            if university is None:
                to_create.append(University(**row))
                continue
            university.label = row["label"]
            university.state = row["state"]
            university.district = row["district"]
            university.location = row["location"]
            university.is_active = True
            to_update.append(university)

        if to_create:
            University.objects.bulk_create(to_create, batch_size=len(to_create))
        if to_update:
            University.objects.bulk_update(
                to_update,
                ["label", "state", "district", "location", "is_active"],
                batch_size=len(to_update),
            )
        return len(to_create), len(to_update)

    @transaction.atomic
    def update_university_types(self, university_type_by_code):
        if not university_type_by_code:
            return
        universities = list(
            University.objects.filter(code__in=university_type_by_code.keys())
        )
        for university in universities:
            university.university_type = university_type_by_code.get(university.code)
        if universities:
            University.objects.bulk_update(
                universities,
                ["university_type"],
                batch_size=len(universities),
            )

    @transaction.atomic
    def upsert_institutions(self, rows):
        existing = {
            item.code: item
            for item in Institution.objects.filter(code__in=[row["code"] for row in rows])
        }
        to_create = []
        to_update = []

        for row in rows:
            institution = existing.get(row["code"])
            if institution is None:
                to_create.append(Institution(**row))
                continue
            institution.label = row["label"]
            institution.institution_type = row["institution_type"]
            institution.university = row["university"]
            institution.state = row["state"]
            institution.district = row["district"]
            institution.location = row["location"]
            institution.college_type = row["college_type"]
            institution.standalone_type = row["standalone_type"]
            institution.management = row["management"]
            institution.university_name = row["university_name"]
            institution.university_type = row["university_type"]
            institution.is_active = True
            to_update.append(institution)

        if to_create:
            Institution.objects.bulk_create(to_create, batch_size=len(to_create))
        if to_update:
            Institution.objects.bulk_update(
                to_update,
                [
                    "label",
                    "institution_type",
                    "university",
                    "state",
                    "district",
                    "location",
                    "college_type",
                    "standalone_type",
                    "management",
                    "university_name",
                    "university_type",
                    "is_active",
                ],
                batch_size=len(to_update),
            )
        return len(to_create), len(to_update)
