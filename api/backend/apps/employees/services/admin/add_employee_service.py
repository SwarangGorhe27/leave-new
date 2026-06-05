"""
Add Employee Service — Admin Side

All DB writes run inside a single atomic transaction.
Each private method owns one domain record.
"""

import logging
import uuid
from copy import deepcopy
from datetime import date
from datetime import datetime, timedelta
import re
from typing import Any, Dict, Optional
from apps.security.models import EmployeeRoleAssignment, Role
from django.contrib.auth import get_user_model
from django.db import IntegrityError, transaction
from rest_framework.exceptions import ValidationError
from apps.leave.helpers import get_employee_for_user
from apps.employees.models.assets import EmployeeAsset
from apps.employees.models.background_verification import EmployeeBackgroundVerification
from apps.employees.models.contact import EmployeeContacts
from apps.employees.models.employee import Employee
from apps.employees.models.employment import EmployeeEmploymentDetails, EmployeeLifecycle
from apps.employees.models.masters.audit_additions import VerificationStatus
from apps.employees.models.masters.performance_training import AssetCategory, AssetCondition
from apps.employees.models.personal import EmployeePersonalDetails
from apps.employees.models.reporting import EmployeeReportingRelationship
from apps.employees.models.statutory import EmployeeStatutoryIds
from apps.employees.serializers.admin.add_employee_serializer import AdminAddEmployeeSerializer
from apps.employees.services.admin import bank_statutory_service
from apps.employees.models.masters.organization import Company
logger = logging.getLogger(__name__)

# Fields stored in Employee.meta_data at creation time.
_EMPLOYEE_META_FIELDS = (
    "employee_number_series",
    "referred_by",
    "onboarding_policy",
    "system_role",
    "allow_employee_to_fill_information",
)

# Statutory ID field mapping: serializer key → model field name.
_STATUTORY_FIELD_MAP: Dict[str, str] = {
    "aadhaar_number": "aadhaar_no",
    "pan_number": "pan",
    "uan_number": "uan",
    "esi_number": "esic_no",
    "passport_number": "passport_no",
}

_ATTENDANCE_KEYS = (
    "working_hours_start",
    "working_hours_end",
    "weekly_off_days",
    "attendance_tracking_mode",
)


class AdminAddEmployeeService:
    """Service layer for the add-employee admin workflow."""

    @staticmethod
    @transaction.atomic
    def create_employee(
        validated_data: Dict[str, Any], created_by=None
    ) -> Dict[str, Any]:
        """Create the full employee record across all related tables.

        Returns the serialized response payload on success.
        Raises ValidationError for duplicate-data or business-rule violations.
        """
        employee_code = validated_data.get("employee_code")
        try:
            # validate payload for common business rules
            AdminAddEmployeeService._validate_input(validated_data)
            employee = AdminAddEmployeeService._create_employee(
                validated_data, created_by
            )
            AdminAddEmployeeService._create_contact_details(employee, validated_data)
            AdminAddEmployeeService._create_personal_details(employee, validated_data)
            AdminAddEmployeeService._create_employment_details(
                employee, validated_data, created_by,
            )
            AdminAddEmployeeService._create_lifecycle(employee, validated_data)
            AdminAddEmployeeService._create_statutory_details(employee, validated_data)
            AdminAddEmployeeService._create_background_check(employee, validated_data)
            AdminAddEmployeeService._create_reporting_relationship(
                employee, validated_data
            )
            AdminAddEmployeeService._create_assets(employee, validated_data)
            # Account access last — needs contacts to already be committed
            AdminAddEmployeeService._create_account_access(
                employee, validated_data, created_by
            )

            logger.info(
                "admin_add_employee_success",
                extra={
                    "employee_id": str(employee.id),
                    "employee_code": employee_code,
                },
            )
            return AdminAddEmployeeService.build_employee_payload(employee.id)

        except ValidationError:
            raise
        except IntegrityError as exc:
            logger.warning(
                "add_employee_integrity_error",
                extra={"employee_code": employee_code},
            )
            raise ValidationError(
                "Employee could not be created due to a duplicate data conflict."
            ) from exc
        except Exception as exc:
            logger.exception(
                "add_employee_unexpected_error",
                extra={"employee_code": employee_code},
            )
            raise ValidationError(
                "Employee could not be created. Please try again or contact support."
            ) from exc

    # -----------------------------------------------------------------------
    # Response payload builder
    # -----------------------------------------------------------------------

    @staticmethod
    def build_employee_payload(employee_id) -> Dict[str, Any]:
        """Fetch and assemble the API response dict for an existing employee."""
        employee = (
            Employee.objects.select_related(
                "company",
                "gender",
                "contacts",
                "employment_details__department",
                "employment_details__designation",
            ).get(id=employee_id)
        )
        contacts = getattr(employee, "contacts", None)
        employment = getattr(employee, "employment_details", None)

        return {
            "employee_id": employee.id,
            "employee_code": employee.employee_code,
            "full_name": employee.full_name,
            "official_email": contacts.official_email if contacts else None,
            "mobile_number": contacts.mobile_no if contacts else None,
            "employee_status": employee.status,
            "company": AdminAddEmployeeService._master_payload(employee.company),
            "department": AdminAddEmployeeService._master_payload(
                getattr(employment, "department", None)
            ),
            "designation": AdminAddEmployeeService._master_payload(
                getattr(employment, "designation", None)
            ),
            "is_draft": bool((employee.meta_data or {}).get("is_draft", False)),
            "is_active": employee.is_active,
        }

    # -----------------------------------------------------------------------
    # Private creation methods
    # -----------------------------------------------------------------------

    @staticmethod
    def _parse_date(val) -> Optional[date]:
        if val is None:
            return None
        if isinstance(val, date):
            return val
        if isinstance(val, str):
            try:
                return date.fromisoformat(val)
            except Exception:
                try:
                    return datetime.fromisoformat(val).date()
                except Exception:
                    return None
        return None

    @staticmethod
    def _validate_input(data: Dict[str, Any]) -> None:
        errors: Dict[str, str] = {}
        # mobile number: ensure last 10 digits are numeric
        mobile = data.get('mobile_number')
        if mobile:
            digits = re.sub(r'\D', '', str(mobile))
            if len(digits) < 10 or len(digits[-10:]) != 10:
                errors['mobile_number'] = 'Mobile number must contain 10 digits'

        emergency = data.get('emergency_contact_number')
        if emergency:
            digits = re.sub(r'\D', '', str(emergency))
            if len(digits) < 10 or len(digits[-10:]) != 10:
                errors['emergency_contact_number'] = 'Emergency contact must contain 10 digits'

        # Aadhaar
        aadhaar = data.get('aadhaar_number')
        if aadhaar:
            digits = re.sub(r'\D', '', str(aadhaar))
            if len(digits) != 12:
                errors['aadhaar_number'] = 'Aadhaar number must be 12 digits'

        # DOB: at least 18 years old
        dob = AdminAddEmployeeService._parse_date(data.get('date_of_birth') or data.get('dob'))
        if dob:
            today = date.today()
            try:
                eighteenth = dob.replace(year=dob.year + 18)
            except Exception:
                # fallback calculation
                eighteenth = dob + timedelta(days=18 * 365)
            if eighteenth > today:
                errors['date_of_birth'] = 'Employee must be at least 18 years old'

        # Joining date within ±15 days
        joining = AdminAddEmployeeService._parse_date(data.get('joining_date') or data.get('date_of_joining') )
        if joining:
            today = date.today()
            low = today - timedelta(days=15)
            high = today + timedelta(days=15)
            if joining < low or joining > high:
                errors['joining_date'] = 'Joining date must be within 15 days of today'

        # Height / weight limits
        height = data.get('height')
        if height:
            try:
                hnum = float(re.sub(r'[^0-9.]', '', str(height)))
                if hnum > 300:
                    errors['height'] = 'Height must be ≤ 300 cm'
            except Exception:
                pass
        weight = data.get('weight')
        if weight:
            try:
                wnum = float(re.sub(r'[^0-9.]', '', str(weight)))
                if wnum > 200:
                    errors['weight'] = 'Weight must be ≤ 200 kg'
            except Exception:
                pass

        # text length limits (150 chars)
        for k, v in list(data.items()):
            if isinstance(v, str) and len(v) > 150:
                errors[k] = 'Maximum length is 150 characters'

        if errors:
            raise ValidationError(errors)

    @staticmethod
    def _create_employee(
        validated_data: Dict[str, Any], created_by
    ) -> Employee:
        actor = AdminAddEmployeeService._resolve_actor(created_by)
        company = AdminAddEmployeeService._resolve_company(
            validated_data, actor, created_by
        )
        validated_data["company"] = company

        meta = {
            "is_draft": validated_data.get("is_draft", False),
            "source": "ADMIN_ADD_EMPLOYEE",
        }
        for key in _EMPLOYEE_META_FIELDS:
            value = validated_data.get(key)
            if value is not None:
                meta[key] = str(value)

        return Employee.objects.create(
            employee_code=validated_data["employee_code"],
            salutation=validated_data.get("salutation"),
            first_name=validated_data["first_name"],
            middle_name=validated_data.get("middle_name"),
            last_name=validated_data["last_name"],
            profile_picture_url=validated_data.get("profile_photo"),
            gender=validated_data["gender"],
            date_of_birth=validated_data["date_of_birth"],
            date_of_joining=validated_data["joining_date"],
            wish_on_date=validated_data["date_of_birth"],
            status=validated_data.get(
                "employee_status", Employee.StatusChoices.ACTIVE
            ),
            company=company,
            is_active=validated_data.get("is_active", True),
            # actor may be None on first-ever employee creation
            created_by=actor if isinstance(actor, Employee) else None,
            updated_by=actor if isinstance(actor, Employee) else None,
            meta_data=meta,
        )

    @staticmethod
    def _create_contact_details(
        employee: Employee, data: Dict[str, Any]
    ) -> None:
        EmployeeContacts.objects.create(
            employee=employee,
            official_email=data["official_email"],
            personal_email=data.get("personal_email"),
            mobile_no=data["mobile_number"],
            alternate_mobile_no=data.get("alternate_mobile_number"),
            emergency_contact_name=data["emergency_contact_name"],
            emergency_contact_phone=data["emergency_contact_number"],
            emergency_contact_relation=data.get("emergency_contact_relationship"),
        )

    @staticmethod
    def _create_personal_details(
        employee: Employee, data: Dict[str, Any]
    ) -> None:
        """Create personal details — only fields that exist on the model."""
        EmployeePersonalDetails.objects.create(
            employee=employee,
            marital_status=data.get("marital_status"),
            blood_group=data.get("blood_group"),
            spouse_name=data.get("spouse_name"),
            father_name=data.get("father_name"),
        )

    @staticmethod
    def _create_background_check(
        employee: Employee, data: Dict[str, Any]
    ) -> None:
        """Create background verification record if any BGC field is provided.

        Uses EmployeeBackgroundVerification model — NOT personal_details.
        verification_status is resolved from VerificationStatus master by code/name.
        """
        bgc_fields = (
            "verification_status",
            "agency_name",
            "verified_by",
            "reference_number",
            "background_remarks",
        )
        if not any(data.get(k) for k in bgc_fields):
            return

        # Resolve verification_status master — accept UUID pk or name/code string
        vs_value = data.get("verification_status")
        vs_instance = None
        if vs_value:
            try:
                vs_instance = VerificationStatus.objects.get(pk=vs_value)
            except (VerificationStatus.DoesNotExist, ValueError):
                vs_instance = (
                    VerificationStatus.objects.filter(
                        name__iexact=str(vs_value)
                    ).first()
                    or VerificationStatus.objects.filter(
                        code__iexact=str(vs_value)
                    ).first()
                )

        if vs_instance is None:
            # BGC data present but status unresolvable — skip silently
            logger.warning(
                "bgc_status_unresolvable",
                extra={
                    "employee_id": str(employee.id),
                    "vs_value": vs_value,
                },
            )
            return

        EmployeeBackgroundVerification.objects.create(
            employee=employee,
            verification_status=vs_instance,
            agency_name=data.get("agency_name"),
            verified_by=data.get("verified_by"),
            reference_number=data.get("reference_number"),
            agency_remarks=data.get("background_remarks"),
        )

    @staticmethod
    def _create_employment_details(
        employee: Employee,
        data: Dict[str, Any],
        created_by,
    ) -> None:
        if not all(
            [data.get("employment_type"), data.get("department"), data.get("designation")]
        ):
            return
        assignment = EmployeeRoleAssignment.objects.create(
            employee=employee,
            role=Role.objects.get(code=data.get("system_role")),
            assigned_by= get_employee_for_user(created_by),
            effective_from=date.today(),
            is_active=True
        )
        actor = AdminAddEmployeeService._resolve_actor(created_by)
        employment = EmployeeEmploymentDetails.objects.create(
            employee=employee,
            employee_type=data["employment_type"],
            department=data["department"],
            designation=data["designation"],
            grade=data.get("grade"),
            office_location=data.get("work_location"),
            shift=data.get("shift"),
            probation_status=AdminAddEmployeeService._probation_label(
                data.get("probation_period")
            ),
            created_by=actor if isinstance(actor, Employee) else None,
            updated_by=actor if isinstance(actor, Employee) else None,
        )

        extras = deepcopy(employment.extra_attributes or {})

        if data.get("business_unit"):
            extras["business_unit_id"] = str(data["business_unit"])
        if data.get("weekly_off"):
            extras["weekly_off"] = data["weekly_off"]
        if data.get("probation_period") is not None:
            extras["probation_period"] = data["probation_period"]

        for key in _ATTENDANCE_KEYS:
            if data.get(key) is not None:
                value = data[key]
                if key in ("working_hours_start", "working_hours_end") and hasattr(
                    value, "isoformat"
                ):
                    extras[key] = value.isoformat()
                else:
                    extras[key] = value

        if data.get("salary_structure") is not None:
            extras["salary_structure"] = str(data["salary_structure"])
        if data.get("basic_salary") is not None:
            extras["basic_salary"] = float(data["basic_salary"])
        if data.get("leave_policy") is not None:
            extras["leave_policy"] = str(data["leave_policy"])
        if data.get("annual_leave_balance") is not None:
            extras["annual_leave_balance"] = int(data["annual_leave_balance"])
        if data.get("sick_leave_balance") is not None:
            extras["sick_leave_balance"] = int(data["sick_leave_balance"])

        employment.extra_attributes = extras
        employment.save()

        if data.get("bank_account"):
            try:
                bank_statutory_service.create_bank_account(
                    str(employee.id), data["bank_account"], created_by
                )
            except Exception as exc:
                raise ValidationError(
                    {"bank_account": "Bank account could not be created or validated."}
                ) from exc

    @staticmethod
    def _create_lifecycle(
        employee: Employee, data: Dict[str, Any]
    ) -> None:
        EmployeeLifecycle.objects.create(
            employee=employee,
            date_of_joining=data["joining_date"],
            probation_start_date=(
                data["joining_date"] if data.get("probation_period") else None
            ),
            date_of_confirmation=data.get("date_of_confirmation"),
        )

    @staticmethod
    def _create_statutory_details(
        employee: Employee, data: Dict[str, Any]
    ) -> None:
        statutory_kwargs = {
            model_field: data[serializer_key]
            for serializer_key, model_field in _STATUTORY_FIELD_MAP.items()
            if data.get(serializer_key)
        }
        if not statutory_kwargs:
            return
        EmployeeStatutoryIds.objects.create(employee=employee, **statutory_kwargs)

    @staticmethod
    def _create_assets(
        employee: Employee, data: Dict[str, Any]
    ) -> None:
        """Create asset records using correct EmployeeAsset field names:
        - assign_date  (not assigned_date)
        - asset_condition  (not condition)
        - serial_no  (not serial_number)
        """
        for index, asset_data in enumerate(data.get("assets") or []):
            try:
                asset_name = asset_data.get("asset_name")
                asset_code = asset_data.get("asset_id") or asset_data.get("asset_code")
                assign_date_raw = asset_data.get("asset_assign_date")

                if not (asset_name and asset_code and assign_date_raw):
                    raise ValidationError(
                        {
                            "assets": f"Asset {index}: asset_name, asset_code, and asset_assign_date are required."
                        }
                    )

                asset_category = AdminAddEmployeeService._resolve_asset_master(
                    AssetCategory, asset_data.get("asset_category")
                )
                asset_condition = AdminAddEmployeeService._resolve_asset_master(
                    AssetCondition, asset_data.get("asset_condition")
                )

                EmployeeAsset.objects.create(
                    employee=employee,
                    asset_name=asset_name,
                    asset_code=asset_code,
                    asset_category=asset_category,
                    asset_condition=asset_condition,
                    serial_no=asset_data.get("serial_number"),
                    assign_date=assign_date_raw,
                    return_date=asset_data.get("asset_return_date"),
                    status=asset_data.get("asset_status", "ASSIGNED"),
                    remarks=asset_data.get("asset_remarks"),
                )

                logger.debug(
                    "asset_created",
                    extra={
                        "employee_id": str(employee.id),
                        "asset_code": asset_code,
                    },
                )
            except ValidationError:
                raise
            except Exception as exc:
                logger.warning(
                    "asset_creation_error",
                    extra={"asset_index": index, "employee_id": str(employee.id)},
                )
                raise ValidationError(
                    {"assets": f"Asset {index}: Could not be created. {str(exc)}"}
                ) from exc

    @staticmethod
    def _create_account_access(
        employee: Employee,
        data: Dict[str, Any],
        created_by,
    ) -> None:
        """Create Django User and link to Employee if username is provided."""
        username = data.get("username")
        if not username:
            return

        # Re-fetch contacts from DB — in-memory relation not guaranteed inside atomic block
        try:
            from apps.employees.models.contact import EmployeeContacts as _EC
            contacts_obj = _EC.objects.filter(employee=employee).first()
            email = contacts_obj.official_email if contacts_obj else data.get("official_email")
        except Exception:
            email = data.get("official_email")

        if not email:
            raise ValidationError(
                {"username": "Official email is required to create a user account."}
            )

        User = get_user_model()
        if User.objects.filter(username=username).exists():
            raise ValidationError({"username": "Username already exists."})

        try:
            if User.objects.filter(email=email).exists():
                user = User.objects.get(email=email)
                logger.info(
                    "account_access_linked_existing_user",
                    extra={"employee_id": str(employee.id), "username": username},
                )
            else:
                user = User.objects.create_user(
                    email=email,
                    password=data.get("temporary_password"),
                    username=username,
                )
                logger.info(
                    "account_access_created_user",
                    extra={"employee_id": str(employee.id), "username": username},
                )

            employee.user = user
            employee.save(update_fields=["user"])
        except Exception as exc:
            logger.warning(
                "account_access_creation_error",
                extra={"employee_id": str(employee.id), "username": username},
            )
            raise ValidationError(
                {"username": "User account could not be created. Please try again."}
            ) from exc

    @staticmethod
    def _create_reporting_relationship(
        employee: Employee, data: Dict[str, Any]
    ) -> None:
        reporting_manager = data.get("reporting_manager")
        if not reporting_manager:
            return

        try:
            EmployeeReportingRelationship.objects.create(
                employee=employee,
                reports_to_employee=reporting_manager,
                relationship_type=EmployeeReportingRelationship.RelationshipType.PRIMARY,
                effective_from=data["joining_date"],
                department=data.get("department"),
                company=data["company"],
                is_active=True,
            )
        except Exception as exc:
            logger.warning(
                "reporting_relationship_creation_error",
                extra={"employee_id": str(employee.id)},
            )
            raise ValidationError(
                {"reporting_manager": "Reporting relationship could not be created."}
            ) from exc

    # -----------------------------------------------------------------------
    # Utility helpers
    # -----------------------------------------------------------------------

    @staticmethod
    def _resolve_company(
        validated_data: Dict[str, Any],
        actor: Optional[Employee],
        created_by,
    ) -> Any:
        # company = validated_data.get("company")
        # if company:
        #     return company
        # if actor:
        #     company = getattr(actor, "company", None)
        # if not company and created_by is not None:
        #     company = getattr(created_by, "company", None)
        # if not company:
        #     raise ValidationError(
        #         {
        #             "company": "Company must be provided or the authenticated user must belong to a company."
        #         }
        #     )
        company = Company.objects.filter(is_active=True).order_by("created_at").first()
        return company

    @staticmethod
    def _resolve_actor(actor) -> Optional[Employee]:
        if actor is None:
            return None
        if isinstance(actor, Employee):
            return actor
        return getattr(actor, "employee_profile", None) or getattr(
            actor, "employee", None
        )

    @staticmethod
    def _resolve_asset_master(model_class, pk_value) -> Optional[Any]:
        if not pk_value:
            return None
        try:
            # AssetCategory / AssetCondition use FullAuditMasterModel which has is_active
            return model_class.objects.get(pk=pk_value, is_active=True)
        except model_class.DoesNotExist:
            return None
        except Exception:
            # Fallback without is_active filter if field doesn't exist
            try:
                return model_class.objects.get(pk=pk_value)
            except model_class.DoesNotExist:
                return None

    @staticmethod
    def _master_payload(obj: Optional[Any]) -> Optional[Dict[str, Any]]:
        if obj is None:
            return None
        return {
            "id": str(obj.id),
            "code": getattr(obj, "code", None),
            "name": (
                getattr(obj, "name", None)
                or getattr(obj, "title", None)
                or getattr(obj, "label", None)
            ),
        }

    @staticmethod
    def _probation_label(probation_period: Optional[int]) -> Optional[str]:
        if probation_period is None:
            return None
        return f"{probation_period} days"


# ---------------------------------------------------------------------------
# Rehire service
# ---------------------------------------------------------------------------

@transaction.atomic
def rehire_employee(validated_data: dict, created_by=None) -> dict:
    """Rehire a former (inactive) employee.

    Creates a new Employee record linked to the former one via
    previous_employee FK. Copies core profile from former employee;
    job/attendance overrides from the request payload take precedence.
    """
    former_id = validated_data["former_employee_id"]

    former = Employee.objects.filter(id=former_id, is_active=False).first()
    if not former:
        raise ValidationError(
            {"former_employee_id": "Former employee not found or is still active."}
        )

    # Resolve company from actor or former employee
    actor = AdminAddEmployeeService._resolve_actor(created_by)
    company = (
        getattr(actor, "company", None)
        or getattr(created_by, "company", None)
        or former.company
    )

    # Resolve job details — use overrides if provided, else copy from former
    former_ed = getattr(former, "employment_details", None)

    employment_type = validated_data.get("employment_type") or getattr(former_ed, "employee_type", None)
    department      = validated_data.get("department")      or getattr(former_ed, "department", None)
    designation     = validated_data.get("designation")     or getattr(former_ed, "designation", None)
    reporting_manager = validated_data.get("reporting_manager") or former.manager
    work_location   = validated_data.get("work_location")   or getattr(former_ed, "office_location", None)
    shift           = validated_data.get("shift")           or getattr(former_ed, "shift", None)

    if not employment_type or not department or not designation:
        raise ValidationError(
            {"employment_type": "employment_type, department and designation are required for rehire (or must exist on former record)."}
        )

    # Generate new employee code
    from apps.employees.models.sequences import EmployeeCodeSequence
    new_code = validated_data.get("new_employee_code")
    if not new_code:
        seq = EmployeeCodeSequence.objects.select_for_update().filter(
            company=company, is_active=True
        ).first()
        if seq:
            seq.last_sequence_no += 1
            seq.save(update_fields=["last_sequence_no"])
            parts = [p for p in [seq.prefix, str(date.today().year) if seq.suffix_format == "YYYY" else None,
                                  str(seq.last_sequence_no).zfill(seq.padding_length)] if p]
            new_code = seq.separator.join(parts)
        else:
            import uuid as _uuid
            new_code = f"EMP-{_uuid.uuid4().hex[:5].upper()}"

    meta = {
        "is_draft": False,
        "source": "ADMIN_REHIRE",
        "rehire_remarks": validated_data.get("rehire_remarks", ""),
        "restore_salary": validated_data.get("restore_salary", False),
        "restore_assets": validated_data.get("restore_assets", False),
        "restore_leaves": validated_data.get("restore_leaves", False),
    }

    new_employee = Employee.objects.create(
        employee_code=new_code,
        salutation=former.salutation,
        first_name=former.first_name,
        middle_name=former.middle_name,
        last_name=former.last_name,
        profile_picture_url=former.profile_picture_url,
        gender=former.gender,
        date_of_birth=former.date_of_birth,
        date_of_joining=validated_data["rehire_date"],
        wish_on_date=former.date_of_birth,
        status=Employee.StatusChoices.ACTIVE,
        company=company,
        manager=reporting_manager,
        is_rehire=True,
        previous_employee=former,
        old_employee_code=former.employee_code,
        is_active=True,
        meta_data=meta,
    )

    # Copy contacts from former
    former_contacts = getattr(former, "contacts", None)
    if former_contacts:
        EmployeeContacts.objects.create(
            employee=new_employee,
            official_email=former_contacts.official_email,
            personal_email=former_contacts.personal_email,
            mobile_no=former_contacts.mobile_no,
            emergency_contact_name=former_contacts.emergency_contact_name,
            emergency_contact_phone=former_contacts.emergency_contact_phone,
            emergency_contact_relation=former_contacts.emergency_contact_relation,
        )

    # Copy personal details from former
    former_personal = getattr(former, "personal_details", None)
    if former_personal:
        EmployeePersonalDetails.objects.create(
            employee=new_employee,
            marital_status=former_personal.marital_status,
            blood_group=former_personal.blood_group,
            spouse_name=former_personal.spouse_name,
            father_name=former_personal.father_name,
        )

    # Employment details
    new_employment = EmployeeEmploymentDetails.objects.create(
        employee=new_employee,
        employee_type=employment_type,
        department=department,
        designation=designation,
        grade=getattr(former_ed, "grade", None),
        office_location=work_location,
        shift=shift,
    )

    # Restore salary/leave extras if requested
    if former_ed and (validated_data.get("restore_salary") or validated_data.get("restore_leaves")):
        extras = deepcopy(former_ed.extra_attributes or {})
        if not validated_data.get("restore_salary"):
            extras.pop("salary_structure", None)
            extras.pop("basic_salary", None)
        if not validated_data.get("restore_leaves"):
            extras.pop("leave_policy", None)
            extras.pop("annual_leave_balance", None)
            extras.pop("sick_leave_balance", None)
        new_employment.extra_attributes = extras
        new_employment.save()

    # Lifecycle
    EmployeeLifecycle.objects.create(
        employee=new_employee,
        date_of_joining=validated_data["rehire_date"],
    )

    # Reporting relationship
    if reporting_manager and hasattr(reporting_manager, "id"):
        EmployeeReportingRelationship.objects.create(
            employee=new_employee,
            reports_to_employee=reporting_manager,
            relationship_type=EmployeeReportingRelationship.RelationshipType.PRIMARY,
            effective_from=validated_data["rehire_date"],
            department=department,
            company=company,
            is_active=True,
        )

    # Restore assets if requested
    if validated_data.get("restore_assets"):
        for old_asset in former.asset_assignments.filter(status="ASSIGNED"):
            EmployeeAsset.objects.create(
                employee=new_employee,
                asset_name=old_asset.asset_name,
                asset_code=old_asset.asset_code,
                asset_category=old_asset.asset_category,
                asset_condition=old_asset.asset_condition,
                serial_no=old_asset.serial_no,
                assign_date=validated_data["rehire_date"],
                status="ASSIGNED",
                remarks=f"Restored from rehire of {former.employee_code}",
            )

    # assignment = EmployeeRoleAssignment.objects.create(
    #             employee=new_employee,
    #             role="passed role",
    #             assigned_by= "admin",
    #             effective_from=date.today(),
    #             is_active=True
    #         )

    # Account access
    username = validated_data.get("username")
    if username:
        User = get_user_model()
        if not User.objects.filter(username=username).exists():
            contacts = getattr(new_employee, "contacts", None)
            email = contacts.official_email if contacts else None
            if email:
                user = User.objects.create_user(
                    email=email,
                    password=validated_data.get("temporary_password"),
                    username=username,
                )
                new_employee.user = user
                new_employee.save(update_fields=["user"])

    logger.info(
        "rehire_employee_success",
        extra={"new_employee_id": str(new_employee.id), "former_employee_id": str(former_id)},
    )
    return AdminAddEmployeeService.build_employee_payload(new_employee.id)


# ---------------------------------------------------------------------------
# Former employees search
# ---------------------------------------------------------------------------

def get_former_employees(company_id: str, search: str = "", department_id: str = "", reason: str = "") -> list:
    """Return inactive employees for the rehire search panel."""
    from django.db.models import Q

    qs = (
        Employee.objects
        .filter(company_id=company_id, is_active=False)
        .select_related(
            "employment_details__department",
            "employment_details__designation",
            "lifecycle",
        )
    )

    if search:
        qs = qs.filter(
            Q(first_name__icontains=search)
            | Q(last_name__icontains=search)
            | Q(employee_code__icontains=search)
            | Q(contacts__official_email__icontains=search)
            | Q(contacts__mobile_no__icontains=search)
        )

    if department_id:
        qs = qs.filter(employment_details__department_id=department_id)

    if reason:
        status_map = {
            "resigned": Employee.StatusChoices.RESIGNED,
            "terminated": Employee.StatusChoices.TERMINATED,
            "relieved": Employee.StatusChoices.RESIGNED,
            "separated": Employee.StatusChoices.TERMINATED,
            "absconded": Employee.StatusChoices.ABSCONDED,
            "retired": Employee.StatusChoices.RETIRED,
        }
        mapped = status_map.get(reason.lower(), reason.upper())
        qs = qs.filter(status=mapped)

    result = []
    for emp in qs[:50]:
        ed = getattr(emp, "employment_details", None)
        lc = getattr(emp, "lifecycle", None)
        result.append({
            "employee_id": str(emp.id),
            "employee_code": emp.employee_code,
            "full_name": emp.full_name,
            "department": getattr(getattr(ed, "department", None), "name", None),
            "designation": getattr(getattr(ed, "designation", None), "title", None),
            "last_working_date": str(lc.relieving_date) if lc and lc.relieving_date else None,
            "separation_reason": emp.status,
            "status": emp.status,
            "profile_picture_url": emp.profile_picture_url,
        })
    return result


# ---------------------------------------------------------------------------
# Bulk import
# ---------------------------------------------------------------------------

def bulk_import_employees(file, created_by=None) -> dict:
    """Parse uploaded Excel/CSV and create employees row by row."""
    import csv
    import io
    from datetime import datetime

    from django.db.models import Q

    from apps.employees.models.masters.employment import EmployeeType
    from apps.employees.models.masters.location import OfficeLocation
    from apps.employees.models.masters.organization import Department, Designation
    from apps.employees.models.masters.personal import Gender
    from apps.employees.models.sequences import EmployeeCodeSequence

    ext = file.name.rsplit(".", 1)[-1].lower()
    rows = []

    def clean_header(value: Any) -> str:
        return " ".join(str(value or "").strip().lower().replace("_", " ").split())

    def clean_value(value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, str):
            value = value.strip()
            return value or None
        return value

    def normalize_date(value: Any) -> Any:
        value = clean_value(value)
        if not value:
            return None
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, datetime):
            return value.date().isoformat()
        value = str(value).strip()
        for date_format in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value, date_format).date().isoformat()
            except ValueError:
                continue
        return value

    def normalize_status(value: Any) -> Any:
        value = clean_value(value)
        if not value:
            return None
        key = str(value).strip().lower().replace(" ", "_").replace("-", "_")
        mapping = {
            "active": Employee.StatusChoices.ACTIVE,
            "inactive": Employee.StatusChoices.RESIGNED,
            "resigned": Employee.StatusChoices.RESIGNED,
            "terminated": Employee.StatusChoices.TERMINATED,
            "on_notice": Employee.StatusChoices.ON_NOTICE,
            "on_leave": Employee.StatusChoices.ON_NOTICE,
            "absconded": Employee.StatusChoices.ABSCONDED,
            "retired": Employee.StatusChoices.RETIRED,
        }
        return mapping.get(key, str(value).strip().upper())

    def error_detail(exc: ValidationError) -> dict:
        detail = getattr(exc, "detail", exc)
        if isinstance(detail, dict):
            return {str(k): " ".join(map(str, v)) if isinstance(v, list) else str(v) for k, v in detail.items()}
        if isinstance(detail, list):
            return {"__all__": " ".join(map(str, detail))}
        return {"__all__": str(detail)}

    def resolve_master(model, value: Any, label: str, company_scoped: bool = False):
        value = clean_value(value)
        if not value:
            raise ValidationError({label: f"{label.replace('_', ' ').title()} is required."})
        text = str(value).strip()
        filters = Q()
        for field in ("code", "label", "name", "title", "prefix"):
            if any(f.name == field for f in model._meta.fields):
                filters |= Q(**{f"{field}__iexact": text})
        queryset = model.objects.filter(is_active=True)
        if company_scoped and any(f.name == "company" for f in model._meta.fields):
            queryset = queryset.filter(company=company)
        match = queryset.filter(filters).first() if filters else None
        if not match:
            raise ValidationError({label: f"{label.replace('_', ' ').title()} not found"})
        return match

    def sequence_exists(value: Any) -> bool:
        value = clean_value(value)
        if not value:
            return False
        return EmployeeCodeSequence.objects.filter(
            company=company,
            is_active=True,
            prefix__iexact=str(value).strip(),
        ).exists()

    if ext == "xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file.read()), read_only=True, data_only=True)
            ws = wb.active
            headers = [
                str(c.value).strip() if c.value else ""
                for c in next(ws.iter_rows(min_row=1, max_row=1))
            ]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(clean_value(cell) is not None for cell in row):
                    rows.append(dict(zip(headers, row)))
        except ImportError:
            raise ValidationError("openpyxl is required for Excel import: pip install openpyxl")
    elif ext == "xls":
        try:
            import xlrd
        except ImportError:
            raise ValidationError("xlrd is required for .xls import.")
        workbook = xlrd.open_workbook(file_contents=file.read())
        sheet = workbook.sheet_by_index(0)
        headers = [str(sheet.cell_value(0, col)).strip() for col in range(sheet.ncols)]
        for row_index in range(1, sheet.nrows):
            values = [sheet.cell_value(row_index, col) for col in range(sheet.ncols)]
            if any(clean_value(cell) is not None for cell in values):
                rows.append(dict(zip(headers, values)))
    elif ext == "csv":
        content = file.read().decode("utf-8-sig")
        rows = [
            row for row in csv.DictReader(io.StringIO(content))
            if any(clean_value(value) is not None for value in row.values())
        ]
    else:
        raise ValidationError("Unsupported file format.")

    if len(rows) > 500:
        raise ValidationError("File exceeds the 500-row limit.")

    field_map = {
        "salutation": "salutation",
        "employee number series": "employee_number_series",
        "employee no": "employee_code",
        "employee code": "employee_code",
        "employee id": "employee_code",
        "emp id": "employee_code",
        "emp code": "employee_code",
        "first name": "first_name",
        "firstname": "first_name",
        "last name": "last_name",
        "lastname": "last_name",
        "email": "official_email",
        "official email": "official_email",
        "mobile number": "mobile_number",
        "mobile": "mobile_number",
        "phone": "mobile_number",
        "date of birth": "date_of_birth",
        "dob": "date_of_birth",
        "aadhaar number": "aadhaar_number",
        "aadhar number": "aadhaar_number",
        "gender": "gender_name",
        "status": "employee_status",
        "date of joining": "joining_date",
        "doj": "joining_date",
        "joining date": "joining_date",
        "department": "department_name",
        "designation": "designation_name",
        "role": "designation_name",
        "designation / role": "designation_name",
        "employment type": "employment_type_name",
        "employee type": "employment_type_name",
        "work location": "work_location_name",
        "location": "work_location_name",
        "system role": "system_role",
    }

    actor = AdminAddEmployeeService._resolve_actor(created_by)
    company = getattr(actor, "company", None) or getattr(created_by, "company", None)
    if not company:
        raise ValidationError({"company": "Admin user has no associated company."})

    results = []
    success_count = failed_count = 0

    for i, row in enumerate(rows, start=2):
        norm = {}
        for k, v in row.items():
            if k:
                key = field_map.get(clean_header(k))
                if key:
                    norm[key] = clean_value(v)

        row_errors = {}
        required_fields = (
            "employee_number_series", "employee_code", "first_name", "last_name",
            "official_email", "mobile_number", "gender_name", "employee_status",
            "joining_date", "department_name", "designation_name",
            "employment_type_name", "work_location_name","system_role"
        )
        for req in required_fields:
            if not norm.get(req):
                row_errors[req] = "This field is required."

        if row_errors:
            results.append({
                "row": i,
                "employee_code": norm.get("employee_code"),
                "status": "failed",
                "errors": row_errors,
            })
            failed_count += 1
            continue

        try:
            with transaction.atomic():
                if not sequence_exists(norm.get("employee_number_series")):
                    raise ValidationError({"employee_number_series": "Employee Number Series not found"})
                if Employee.objects.filter(employee_code__iexact=norm["employee_code"]).exists():
                    raise ValidationError({"employee_code": "Employee Code already exists"})
                if EmployeeContacts.objects.filter(official_email__iexact=norm["official_email"]).exists():
                    raise ValidationError({"official_email": "Email already exists"})
                if EmployeeContacts.objects.filter(mobile_no__iexact=str(norm["mobile_number"])).exists():
                    raise ValidationError({"mobile_number": "Mobile already exists"})
                if norm.get("aadhaar_number") and EmployeeStatutoryIds.objects.filter(
                    aadhaar_no__iexact=str(norm["aadhaar_number"])
                ).exists():
                    raise ValidationError({"aadhaar_number": "Aadhaar already exists"})
                role = Role.objects.filter(
                code__iexact=norm["system_role"],
                is_active=True
                        ).first()
                print(role.code)
                if not role:
                    raise ValidationError({
                        "system_role": f"Role '{norm['system_role']}' not found"
                    })
                payload = {
                    "salutation": norm.get("salutation"),
                    "employee_number_series": norm["employee_number_series"],
                    "employee_code": norm["employee_code"],
                    "first_name": norm["first_name"],
                    "last_name": norm["last_name"],
                    "date_of_birth": normalize_date(norm.get("date_of_birth")),
                    "aadhaar_number": norm.get("aadhaar_number"),
                    "gender": resolve_master(Gender, norm.get("gender_name"), "gender"),
                    "joining_date": normalize_date(norm["joining_date"]),
                    "employee_status": normalize_status(norm["employee_status"]),
                    "official_email": norm["official_email"],
                    "mobile_number": str(norm["mobile_number"]),
                    "emergency_contact_name": "",
                    "emergency_contact_number": "",
                    "employment_type": resolve_master(
                        EmployeeType,
                        norm.get("employment_type_name"),
                        "employment_type",
                    ),
                    "system_role": role.code,
                    "department": resolve_master(Department, norm.get("department_name"), "department", True),
                    "designation": resolve_master(Designation, norm.get("designation_name"), "designation", True),
                    "work_location": resolve_master(OfficeLocation, norm.get("work_location_name"), "work_location"),
                    "company": company,
                    "is_draft": False,
                    "is_active": True,
                }

                serializer = AdminAddEmployeeSerializer(data=payload)
                if not serializer.is_valid():
                    raise ValidationError(serializer.errors)
                result = AdminAddEmployeeService.create_employee(
                    serializer.validated_data,
                    created_by,
                )
            results.append({
                "row": i,
                "status": "success",
                "employee_code": result["employee_code"],
            })
            success_count += 1
            logger.info(
                "bulk_import_row_success",
                extra={"row": i, "employee_code": result["employee_code"]},
            )
        except ValidationError as exc:
            results.append({
                "row": i,
                "status": "failed",
                "employee_code": norm.get("employee_code"),
                "errors": error_detail(exc),
            })
            failed_count += 1
            logger.warning(
                "bulk_import_row_validation_error",
                extra={"row": i, "error": str(exc)},
            )
        except Exception as exc:
            results.append({
                "row": i,
                "status": "failed",
                "employee_code": norm.get("employee_code"),
                "errors": {"__all__": "Employee could not be imported. Please review this row."},
            })
            failed_count += 1
            logger.exception(
                "bulk_import_row_unexpected_error",
                extra={"row": i},
            )

    logger.info(
        "bulk_import_completed",
        extra={
            "total_rows": len(rows),
            "success_count": success_count,
            "failed_count": failed_count,
            "company_id": str(company.id),
        },
    )

    return {
        "total_rows": len(rows),
        "success_count": success_count,
        "failed_count": failed_count,
        "results": results,
    }


def generate_bulk_import_template():
    """Return an openpyxl Workbook with the standard import template."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ValidationError("openpyxl is required: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Import"

    headers = [
        "Salutation",
        "Employee Number Series",
        "Employee No",
        "First Name",
        "Last Name",
        "Email",
        "Mobile Number",
        "Date of Birth",
        "Aadhaar Number",
        "Gender",
        "Status",
        "Date of Joining",
        "Department",
        "Designation / Role",
        "Employment Type",
        "Work Location",
        "System Role",
    ]
    ws.append(headers)

    fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.column_letter  # trigger dimension calc

    return wb


# ---------------------------------------------------------------------------
# Draft save / get  (in-memory; swap with DB model for production)
# ---------------------------------------------------------------------------

_DRAFT_STORE: dict = {}


def save_draft(validated_data: dict, user_id: str) -> dict:
    import uuid as _uuid
    from django.utils import timezone

    draft_id = str(validated_data.get("draft_id") or _uuid.uuid4())
    _DRAFT_STORE[draft_id] = {
        "draft_id": draft_id,
        "draft_type": validated_data.get("draft_type", "new"),
        "draft_data": validated_data.get("draft_data", {}),
        "saved_at": timezone.now().isoformat(),
        "user_id": user_id,
    }
    return {
        "draft_id": draft_id,
        "draft_type": _DRAFT_STORE[draft_id]["draft_type"],
        "saved_at": _DRAFT_STORE[draft_id]["saved_at"],
    }


def get_draft(draft_id: str, user_id: str) -> dict:
    draft = _DRAFT_STORE.get(str(draft_id))
    if not draft or draft.get("user_id") != user_id:
        raise ValidationError({"draft_id": "Draft not found."})
    return draft
